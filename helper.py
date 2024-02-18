import os
import openpyxl
import datetime
from flask import make_response
import shutil

#ユーザー定義フォーマットを追加
currentJPY = '"¥"#,##0'

def writeExcel(exportJobList, clientName):
    #今日の日付を取得して、請求書の発行日として入力
    nowDate = datetime.date.today()

    #テンプレートファイルのパス
    templatefile = 'invoice-template.xlsx'
    #出力する請求書のファイル名（先頭に本日の日付を入れる）
    outputfile = str(nowDate) + '-invoice.xlsx'

    #テンプレートファイルをコピーして、出力ファイルを新規作成
    shutil.copy(templatefile, outputfile)

    #出力ファイルを読み込み、書き込むワークシートを指定
    workbook = openpyxl.load_workbook(outputfile)
    worksheet = workbook.get_sheet_by_name(u'シート1')

    #引数のexportJobListから、各項目と金額をoutputfileに記入
    for i in range(len(exportJobList)):
        #作業内容
        worksheet.cell(row = i + 12, column = 1).value = exportJobList[i][0]
        #作業した日付
        worksheet.cell(row = i + 12, column = 2).value = exportJobList[i][2]
        #税率
        tax_rate = int(exportJobList[i][3] * 100)
        worksheet.cell(row = i + 12, column = 4).value = str(tax_rate) + "%"
        #金額を「¥0,000」形式で入力
        worksheet.cell(row = i + 12, column = 5).number_format = currentJPY
        worksheet.cell(row = i + 12, column = 5).value = exportJobList[i][1]
    
    #請求書発行日をE3セルに入力
    worksheet["E3"].value = str(nowDate)
    #クライアント名をA3セルに入力
    worksheet["A3"].value = str(clientName)
    #小計を「¥0,000」形式でE41セルに入力
    worksheet["E41"].number_format = currentJPY
    worksheet["E41"].value = "=SUM(E12:E40)"

    #税率ごとの合計金額を計算
    zeroTaxTotal = 0
    lightTaxTotal = 0
    normalTaxTotal = 0
    #worksheetのD列を走査して、税率が8%のセルを探す
    for i in range(12, 41):
        if worksheet.cell(row = i, column = 4).value == "8%":
            lightTaxTotal += worksheet.cell(row = i, column = 5).value
        elif worksheet.cell(row = i, column = 4).value == "10%":
            normalTaxTotal += worksheet.cell(row = i, column = 5).value
        elif worksheet.cell(row = i, column = 4).value == "0%":
            zeroTaxTotal += worksheet.cell(row = i, column = 5).value

    #8%税率対象の合計金額をB41セルに入力
    worksheet["B41"].number_format = currentJPY
    worksheet["B41"].value = lightTaxTotal

    #8%税率の税額をC41セルに入力
    worksheet["C41"].number_format = currentJPY
    worksheet["C41"].value = int(lightTaxTotal * 0.08)

    #10%税率対象の合計金額をB42セルに入力
    worksheet["B42"].number_format = currentJPY
    worksheet["B42"].value = normalTaxTotal

    #10%税率の税額をC42セルに入力
    worksheet["C42"].number_format = currentJPY
    worksheet["C42"].value = int(normalTaxTotal * 0.1)

    #0%税率対象の合計金額をB43セルに入力
    worksheet["B43"].number_format = currentJPY
    worksheet["B43"].value = zeroTaxTotal
    
    #0%税率の税額をC43セルに入力
    worksheet["C43"].number_format = currentJPY
    worksheet["C43"].value = int(zeroTaxTotal * 0.0)
    
    #消費税合計金額をB44セルに入力
    worksheet["B44"].number_format = currentJPY
    worksheet["B44"].value = "=SUM(B41:B43)"

    #消費税込み合計金額をE42セルに入力
    worksheet["E42"].number_format = currentJPY
    worksheet["E42"].value = "=SUM(E41,B44)"

    #ファイル上部のB7セルにも合計金額を入力
    worksheet["B7"].number_format = currentJPY
    worksheet["B7"].value = "=SUM(E41, B44)"

    #outputfileを保存する
    workbook.save(outputfile)
    workbook.close()

    #ブラウザからExcelをダウンロードする
    XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spredsheetml.sheet"

    response = make_response()

    workbook = open(outputfile, "rb")
    response.data = workbook.read()
    workbook.close()

    response.headers["Content-Disposition"] = "attatchment; filename=" + str(nowDate) + "-invoice.xlsx"
    response.mimetype = XLSX_MIMETYPE
    #出力したらoutputfileを削除
    os.remove(outputfile)
    return response

    